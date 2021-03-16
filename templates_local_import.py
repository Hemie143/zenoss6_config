import zenAPI.zenApiLib
import re
import time
import argparse
import openpyxl

from templates_tools import get_template_uid, get_parent_template

def create_local_template(router, device_uid, template_name):
    # template_uid = get_higher_template(router, device_uid, template_name)
    # if not template_uid:
    #     return
    template_uid = '{}/{}'.format(device_uid, template_name)
    response = router.callMethod('getInfo', uid=template_uid)
    if response['result']['success']:
        # print('Local Template already exists')
        return template_uid
    else:
        response = router.callMethod('makeLocalRRDTemplate', uid=device_uid, templateName=template_name)
        # print(response)
        return response['result']['tplUid']
    return


def set_datasource_prop(router, template_uid, action, object_name, key, value):
    # Check datasource presence
    # Edit property
    ds_uid = '{}/datasources/{}'.format(template_uid, object_name)
    result = router.callMethod('getInfo', uid=ds_uid)['result']
    if not result['success']:
        print('Datasource not found')
        response = router.callMethod('addDataSource', templateUid=template_uid, name=object_name, type=type)
        print(response)
    else:
        print('Datasource already exists')

    return

def process_templates(device_router, template_router, data):

    rows = data.max_row
    for r, (template_name, device_class, device, component, action, key, value) in enumerate(data.iter_rows(min_row=2, max_col=7)):

        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Templates')

        # TODO: check if template_uid is None
        t_result = template_router.callMethod('getInfo', uid=template_uid)['result']
        t_present = t_result['success'] and t_result['data']['uid'] == template_uid

        msg = 'Templates Row {}/{}: Template unchanged'.format(r + 2, rows)

        if action.value == 'ADD':
            if t_present:
                pass
            else:
                if device.value:
                    template_uid = create_local_template(template_router, parent_uid, template_name.value)
                else:
                    # Device class
                    # TODO: test this again, refactor the names, they are confusing
                    template_source = get_parent_template(template_router, template_uid)
                    if template_source:
                        t_result = template_router.callMethod('copyTemplate', uid=template_source,
                                                              targetUid=parent_uid)['result']
                    else:
                        t_result = template_router.callMethod('addTemplate', id=template_name.value,
                                                              targetUid=parent_uid)['result']
                    if t_result['success']:
                        template_uid = 'OK'
                    else:
                        template_uid = None
                if template_uid:
                    msg = 'Templates Row {}/{}: Template added'.format(r + 2, rows)
                else:
                    msg = 'Templates Row {}/{}: Template failed to add: {}'.format(r + 2, rows, template_uid)
        elif action.value == 'DELETE':
            msg = 'Templates Row {}/{}: DELETE action to implement'.format(r + 2, rows)
        else:
            msg = 'Templates Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value)

        if key.value:
            if action.value == 'ADD':
                current_value = t_result['data'].get(key.value, u'')
                if isinstance(current_value, unicode):
                    new_value = unicode(str(value.value), 'utf-8')
                elif isinstance(current_value, bool):
                    if value.value.lower() in ['false', '0']:
                        new_value = False
                    elif value.value.lower() in ['true', '1']:
                        new_value = True
                    else:
                        msg = 'Templates Row {}/{}: Value not recognized'.format(r + 2, rows)
                elif isinstance(current_value, int):
                    new_value = int(value.value)
                else:
                    msg = 'Templates Row {}/{}: ERROR: Datasource value unknown: {}'.format(r + 2, rows,
                                                                                           type(current_value))
                    print('-- Templates type current: {}'.format(type(current_value)))
                    print('-- Templates type new    : {}'.format(type(str(value.value))))
                    print('-- Templates value current: {}'.format(current_value))
                    print('-- Templates value new    : {}'.format((str(value.value))))
                    print('-- Templates test current    : {}'.format((isinstance(current_value, unicode))))
                    continue

                if current_value == new_value:
                    msg = 'Templates Row {}/{}: Template unchanged'.format(r + 2, rows)
                else:
                    data = {u'uid': t_result['data']['uid'], key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        msg = 'Templates Row {}/{}: Template edited'.format(r + 2, rows)
                    else:
                        msg = 'Templates Row {}/{}: Template failed'.format(r + 2, rows)
                        print(result)
            elif action.value == 'DELETE':
                msg = 'Templates Row {}/{}: DELETE action to implement'.format(r + 2, rows)
            else:
                msg = 'Templates Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value)
        print(msg)
    return

def process_datasources(device_router, template_router, data):

    rows = data.max_row
    for r, (template_name, device_class, device, component, action, ds_name, ds_type, key, value) in enumerate(data.iter_rows(min_row=2, max_col=9)):

        # Find device UID
        if not template_name.value:
            continue

        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Datasources')

        # Find template UID and check presence
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Datasources Row {}/{}: Template not found'.format(r + 2, rows, template_name))
            continue

        # Find datasource UID and check presence
        ds_uid = '{}/datasources/{}'.format(template_uid, ds_name.value)
        ds_result = template_router.callMethod('getInfo', uid=ds_uid)['result']
        if ds_result['success']:
            # Datasource is present
            if action.value == 'ADD':
                # print('Datasources Row {}: Datasource present'.format(r + 2))
                pass
            elif action.value == 'DELETE':
                result = template_router.callMethod('deleteDataSource', uid=ds_uid)['result']
                if result['success']:
                    print('Datasources Row {}/{}: Datasource deleted'.format(r + 2, rows))
                else:
                    print('Datasources Row {}/{}: ERROR: Datasource delete failed'.format(r + 2, rows))
            else:
                print('Datasources Row {}/{}: ERROR: action unknown: {}'.format(r + 2, rows, action.value))
        else:
            # Datasource is not present
            if action.value == 'ADD':
                response = template_router.callMethod('addDataSource', templateUid=template_uid, name=ds_name.value, type=ds_type.value)
                ds_result = response['result']
                if ds_result['success']:
                    ds_result = template_router.callMethod('getInfo', uid=ds_uid)['result']
                    print('Datasources Row {}/{}: Datasource added'.format(r + 2, rows))
                else:
                    print('Datasources Row {}/{}: Datasource failed to add: {}'.format(r + 2, rows, ds_result['msg']))
                    continue
            elif action.value == 'DELETE':
                print('Datasources Row {}/{}: Datasource already deleted'.format(r + 2, rows))
            else:
                print('Datasources Row {}/{}: ERROR: action unknown: {}'.format(r + 2, rows, action.value))

        # Apply datasource edit
        if key.value:
            if action.value == 'ADD':
                # print(ds_result)
                if key.value in ['commandTemplate', 'oid']:
                    current_value = ds_result['data'].get('source', u'')
                    if current_value.endswith(' over SSH'):
                        current_value = current_value[:-9]
                elif key.value == 'usessh':
                    source = ds_result['data'].get('source', u'')
                    current_value = source.endswith(' over SSH')
                else:
                    current_value = ds_result['data'].get(key.value, u'')
                    # print(ds_result)
                    # print('--current_value: **{}/{}**'.format(current_value, type(current_value)))
                if isinstance(current_value, unicode):
                    new_value = unicode(str(value.value), 'utf-8')
                elif isinstance(current_value, bool):
                    if value.value.lower() in ['false', '0']:
                        new_value = False
                    elif value.value.lower() in ['true', '1']:
                        new_value = True
                    else:
                        print('Datasources Row {}/{}: Value not recognized'.format(r + 2, rows))
                elif isinstance(current_value, int):
                    try:
                        new_value = int(value.value)
                    except ValueError:
                        pass
                else:
                    print('Datasources Row {}/{}: ERROR: Datasource value unknown: {}'.format(r + 2, rows,
                                                                                           type(current_value)))
                    print('-- Datasource type current: {}'.format(type(current_value)))
                    print('-- Datasource type new    : {}'.format(type(str(value.value))))
                    print('-- Datasource value current: {}'.format(current_value))
                    print('-- Datasource value new    : {}'.format((str(value.value))))
                    print('-- Datasource test current    : {}'.format((isinstance(current_value, unicode))))
                    continue

                if current_value == new_value:
                    print('Datasources Row {}/{}: Datasource unchanged'.format(r + 2, rows))
                else:
                    # print('new_value    : **{}/{}**'.format(new_value, type(new_value)))
                    # print('current_value: **{}/{}**'.format(current_value, type(current_value)))
                    data = {u'uid': ds_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        print('Datasources Row {}/{}: Datasource edited'.format(r + 2, rows))
                    else:
                        print('Datasources Row {}/{}: Datasource failed'.format(r + 2, rows))
                        print(result)
            elif action.value == 'DELETE':
                print('Datasources Row {}/{}: DELETE action to implement'.format(r + 2, rows))
            else:
                print('Datasources Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))
    return

def process_datapoints(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device_class, device, component, action, ds_name, dp_name, key, value) in enumerate(data.iter_rows(min_row=2, max_col=9)):

        if not template_name.value:
            continue

        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Datapoints')

        # Find datasource UID and check presence
        ds_uid = '{}/datasources/{}'.format(template_uid, ds_name.value)
        result = template_router.callMethod('getInfo', uid=ds_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                print('Datapoints Row {}/{}: ERROR: Datasource not found'.format(r + 2, rows, ds_name))
                continue
            elif action.value == 'DELETE':
                print('Datapoints Row {}/{}: ERROR: DELETE action to implement'.format(r + 2, rows))
            else:
                print('Datapoints Row {}/{}: ERROR: action unknown: {}'.format(r + 2, rows, action.value))

        # Create datapoint
        dp_uid = '{}/datapoints/{}'.format(ds_uid, dp_name.value)
        dp_result = template_router.callMethod('getInfo', uid=dp_uid)['result']
        dp_present = dp_result['success']
        # print(result)
        if action.value == 'ADD':
            if dp_present:
                # Datapoint is present
                print('Datapoints Row {}/{}: Datapoint unchanged'.format(r + 2, rows))
            else:
                # Datapoint doesn't exist yet
                result = template_router.callMethod('addDataPoint', dataSourceUid=ds_uid, name=dp_name.value)['result']
                if result['success']:
                    print('Datapoints Row {}/{}: Datapoint added: {}'.format(r + 2, rows, dp_name.value))
                else:
                    print('Datapoints Row {}/{}: ERROR: Datapoint failed: {}'.format(r + 2, rows, dp_name.value))
        elif action.value == 'DELETE':
            if dp_present and not key.value:
                result = template_router.callMethod('deleteDataPoint', uid=dp_uid)['result']
                if result['success']:
                    print('Datapoints Row {}/{}: Datapoint deleted'.format(r + 2, rows))
                else:
                    print('Datapoints Row {}/{}: ERROR: Datapoint delete failed'.format(r + 2, rows))
                continue
            else:
                print('Datapoints Row {}/{}: Datapoint already deleted'.format(r + 2, rows))
        else:
            print('Datapoints Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))
            continue

        # Edit datapoint
        if key.value:
            if action.value == 'ADD':
                current_value = dp_result['data'].get(key.value, u'')
                if isinstance(current_value, unicode):
                    new_value = unicode(str(value.value), 'utf-8')
                elif isinstance(current_value, bool):
                    if value.value.lower() in ['false', '0']:
                        new_value = False
                    elif value.value.lower() in ['true', '1']:
                        new_value = True
                    else:
                        print('Datapoints Row {}/{}: Value not recognized'.format(r + 2, rows))
                elif isinstance(current_value, int):
                    try:
                        new_value = int(value.value)
                    except ValueError:
                        pass
                elif current_value is None:
                    new_value = value.value
                else:
                    print('Datapoints Row {}/{}: ERROR: Datasource value unknown: {}'.format(r + 2, rows,
                                                                                           type(current_value)))
                    print('-- Datapoints type current: {}'.format(type(current_value)))
                    print('-- Datapoints type new    : {}'.format(type(str(value.value))))
                    print('-- Datapoints value current: {}'.format(current_value))
                    print('-- Datapoints value new    : {}'.format((str(value.value))))
                    print('-- Datapoints test current    : {}'.format((isinstance(current_value, unicode))))
                    continue

                if current_value == new_value:
                    print('Datapoints Row {}/{}: Datapoint unchanged'.format(r + 2, rows))
                else:
                    # print('new_value    : **{}/{}**'.format(new_value, type(new_value)))
                    # print('current_value: **{}/{}**'.format(current_value, type(current_value)))
                    data = {u'uid': dp_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        print('Datapoints Row {}/{}: Datasource edited'.format(r + 2, rows))
                    else:
                        print('Datapoints Row {}/{}: Datasource failed'.format(r + 2, rows))
                        print(result)
            elif action.value == 'DELETE':
                print('Datapoints Row {}/{}: DELETE action to implement'.format(r + 2, rows))
            else:
                print('Datapoints Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))

    return

def process_thresholds(device_router, template_router, data):
    rows = data.max_row
    for r, (template_name, device_class, device, component, action, th_name, th_type, dp_list, key,
            value) in enumerate(data.iter_rows(min_row=2, max_col=10)):
        # TODO: change the way that datapoints are added or deleted. It would be easier to have the datasource
        #  separated from the datapoint and to have a datapoint per row instead of having to manage a list

        if not template_name.value:
            continue

        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Thresholds')
        if not template_uid:
            print('Thresholds Row {}/{}: Template not found'.format(r + 2, rows))
            continue

        # Check presence of threshold
        th_uid = '{}/thresholds/{}'.format(template_uid, th_name.value)
        th_result = template_router.callMethod('getInfo', uid=th_uid)['result']
        th_present = th_result['success']

        # Fetch threshold's datapoints
        if th_present:
            # DS Names have the format <ds_name>_<dp_name> (underscore)
            dsnames = th_result['data']['dsnames']
            # print(sorted(dsnames))
            # print(th_result['data'])
            th_datapoints = []
            for d in dsnames:
                # dlist = d.split('_')
                # th_datapoints.append(u'{}.{}'.format('_'.join(dlist[:-1]), dlist[-1]))
                # Not exactly accurate, but best solution for now
                th_datapoints.append(d.replace('_','.'))
            th_datapoints = sorted(th_datapoints)
        else:
            th_datapoints = []

        datapoints = []
        if dp_list.value:
            datapoints = sorted([d.strip() for d in dp_list.value.split(',')])
            datapoints_ = [d.replace('_', '.') for d in datapoints]
        # Datapoints have the format <ds_name>.<dp_name>  (dot)
        all_datapoints = template_router.callMethod('getDataPoints', uid=template_uid)['result']['data']

        if action.value == 'ADD':
            if th_present and th_datapoints == datapoints_:
                # Datapoint is present
                if not key.value:
                    print('Thresholds Row {}/{}: Threshold unchanged: {}'.format(r + 2, rows, th_name.value))
            else:
                # Datapoint doesn't exist yet
                # Find datapoint UIDs
                dp_uids = []    # TODO: compute UIDs of each datapoint
                for dp in datapoints:
                    for d in all_datapoints:
                        if d['name'] == dp:
                            dp_uids.append(d['uid'])
                if len(dp_uids) != len(datapoints):
                    print('Thresholds Row {}/{}: Datapoints not found: {}'.format(r + 2, rows, [d['name'] for d in all_datapoints]))
                    continue

                data = {u'uid': template_uid, u'thresholdType': th_type.value, u'thresholdId': th_name.value, u'dataPoints': dp_uids}
                result = template_router.callMethod('addThreshold', **data)['result']
                if result['success']:
                    print('Thresholds Row {}/{}: Threshold added: {}'.format(r + 2, rows, th_name.value))
                else:
                    print('Thresholds Row {}/{}: ERROR: Threshold creation failed: {}: {}'.format(r + 2, rows, th_name.value, result['msg']))
                    print(result)
                    continue
                th_result = template_router.callMethod('getInfo', uid=th_uid)['result']
        elif action.value == 'DELETE':
            if th_present and not key.value:
                result = template_router.callMethod('removeThreshold', uid=th_uid)['result']
                if result['success']:
                    print('Thresholds Row {}/{}: Threshold deleted'.format(r + 2, rows))
                else:
                    print('Thresholds Row {}/{}: Threshold failed to delete'.format(r + 2, rows))
            else:
                print('Thresholds Row {}/{}: Threshold already deleted'.format(r + 2, rows))
        else:
            print('Thresholds Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))
            continue

        # Edit threshold
        if key.value:
            # print(th_result)
            current_value = th_result['data'][key.value]
            if isinstance(current_value, unicode):
                new_value = unicode(str(value.value), 'utf-8')
            elif isinstance(current_value, bool) or key.value == 'enabled':
                new_value = bool(value.value)
                if value.value.lower() in ['false', '0']:
                    new_value = False
                elif value.value.lower() in ['true', '1']:
                    new_value = True
                else:
                    print('Thresholds Row {}/{}: Unrecognized value: {}'.format(r + 2, rows, value.value))
            elif isinstance(current_value, int):
                new_value = int(value.value)
            else:
                print('Thresholds Row {}/{}: ERROR: Graph value unknown: {}'.format(r + 2, rows, type(current_value)))
                print('-- Thresholds type current: {}'.format(type(current_value)))
                print('-- Thresholds type new    : {}'.format(type(str(value.value))))
                print('-- Thresholds value current: {}'.format(current_value))
                print('-- Thresholds value new    : {}'.format((str(value.value))))
                print('-- Thresholds test current    : {}'.format((isinstance(current_value, unicode))))
                continue

            if action.value == 'ADD':
                if current_value == new_value:
                    print('Thresholds Row {}/{}: Threshold unchanged'.format(r + 2, rows))
                else:
                    data = {u'uid': th_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        print('Thresholds Row {}/{}: Threshold edited'.format(r + 2, rows))
                    else:
                        print('Thresholds Row {}/{} Threshold failed'.format(r + 2, rows))
            elif action.value == 'DELETE':
                print('Thresholds Row {}/{}: DELETE action to implement'.format(r + 2, rows))
            else:
                print('Thresholds Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value))
    return

def process_graphs(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''
    rows = data.max_row
    for r, (template_name, device_class, device, component, action, gr_name, key, value) in enumerate(data.iter_rows(min_row=2, max_col=8)):

        if not template_name.value:
            continue

        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Graphs')

        # Find graph UID and check presence
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Graphs Row {}: Template not found'.format(r + 2, template_name))
            continue
        gr_uid = '{}/graphDefs/{}'.format(template_uid, gr_name.value)

        gr_result = template_router.callMethod('getInfo', uid=gr_uid)['result']
        gr_present = (gr_result['success']) and (gr_result['data']['inspector_type'] == 'GraphDefinition')

        if action.value == 'ADD':
            if gr_present:
                # print('Graphs Row {}: Graph added'.format(r + 2))
                msg = 'Graphs Row {}/{}: Graph already present'.format(r + 2, rows)
            else:
                response = template_router.callMethod('addGraphDefinition', templateUid=template_uid, graphDefinitionId=gr_name.value)
                gr_result = response['result']
                if gr_result['success']:
                    gr_result = template_router.callMethod('getInfo', uid=gr_uid)['result']
                    msg = 'Graphs Row {}: Graph added'.format(r + 2)
                else:
                    msg = 'Graphs Row {}: ERROR: Graph not added'.format(r + 2)
        elif action.value == 'DELETE':
            if not key.value:
                if gr_present:
                    result = template_router.callMethod('deleteGraphDefinition', uid=gr_uid)
                    if result['result']['success']:
                        msg = 'Graphs Row {}/{}: Graph deleted'.format(r + 2, rows)
                    else:
                        msg = 'Graphs Row {}/{}: Graphed failed to delete'.format(r + 2, rows)
                else:
                    msg = 'Graphs Row {}/{}: Graph already deleted'.format(r + 2, rows)
        else:
            msg = 'Graphs Row {}: action unknown: {}'.format(r + 2, action.value)

        # Apply graph edit
        if key.value:
            # print(gr_result)
            current_value = gr_result['data'][key.value]
            if isinstance(current_value, unicode):
                new_value = unicode(str(value.value), 'utf-8')
            elif isinstance(current_value, bool):
                if value.value.lower() in ['false', '0']:
                    new_value = False
                elif value.value.lower() in ['true', '1']:
                    new_value = True
                else:
                    print('Datapoints Row {}/{}: Value not recognized'.format(r + 2, rows))
            elif isinstance(current_value, int):
                new_value = int(value.value)
            else:
                print('Graphs Row {}: ERROR: Graph value unknown: {}'.format(r + 2,type(current_value)))
                print('-- Graphs type current: {}'.format(type(current_value)))
                print('-- Graphs type new    : {}'.format(type(str(value.value))))
                print('-- Graphs value current: {}'.format(current_value))
                print('-- Graphs value new    : {}'.format((str(value.value))))
                print('-- Graphs test current    : {}'.format((isinstance(current_value, unicode))))
                continue
            if action.value == 'ADD':
                if current_value == new_value:
                    msg = 'Graphs Row {}/{}: Graph unchanged'.format(r + 2, rows)
                else:
                    data = {u'uid': gr_uid, key.value: new_value}
                    result = template_router.callMethod('setInfo', **data)['result']
                    if result['success']:
                        msg = 'Graphs Row {}/{}: Graph edited'.format(r + 2, rows)
                    else:
                        msg = 'Graphs Row {}/{}: Graph failed'.format(r + 2, rows)
            elif action.value == 'DELETE':
                msg = 'Graphs Row {}/{}: DELETE action to implement'.format(r + 2, rows)
            else:
                msg = 'Graphs Row {}/{}: action unknown: {}'.format(r + 2, rows, action.value)
        print(msg)
    return

def process_graphpoints(device_router, template_router, data):
    '''
    for r, this_row in enumerate(data.iter_rows(min_row=2)):
        print(this_row)
        print(this_row[2].value)
    '''

    # print('Processing Graphpoints==============================')
    rows = data.max_row
    for r, (template_name, device_class, device, component, action, gr_name, gp_name, ds_name, dp_name, key, value) in \
            enumerate(data.iter_rows(min_row=2, max_col=11)):

        if not template_name.value:
            continue

        # Find template UID
        parent_uid, template_uid = get_template_uid(device_router,
                                                    template_name.value,
                                                    device_class.value,
                                                    device.value,
                                                    component.value,
                                                    r+2,
                                                    'Graphpoints')

        # Check template presence
        result = template_router.callMethod('getInfo', uid=template_uid)['result']
        if not result['success']:
            print('Graphpoints Row {}: Template not found'.format(r + 2, template_name))
            continue

        # Find graph UID and check presence
        gr_uid = '{}/graphDefs/{}'.format(template_uid, gr_name.value)
        result = template_router.callMethod('getInfo', uid=gr_uid)['result']
        if not result['success']:
            if action.value == 'ADD':
                print('Graphpoints Row {}: ERROR: Graph not found'.format(r + 2, gr_name.value))
                continue
            elif action.value == 'DELETE':
                print('Graphpoints Row {/{}: Already deleted'.format(r + 2, rows))
            else:
                print('Graphpoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))

        # Find datapoint or threshold UID and check presence
        point_name = ''
        dp_uid = ''
        th_uid = ''
        if ds_name.value:
            # This should be a datapoint
            # /zport/dmd/Devices/Server/Linux/APP/Prod/devices/prb-app-l02.in.credoc.be/Tomcat_JMX/datasources/Auction_ClassLoading_LoadedClassCount/datapoints/LoadedClassCount
            dp_uid = '{}/datasources/{}/datapoints/{}'.format(template_uid, ds_name.value, dp_name.value)
            dp_result = template_router.callMethod('getInfo', uid=dp_uid)['result']
            if not dp_result['success']:
                if action.value == 'ADD':
                    print('Graphpoints Row {}: ERROR: Datapoint not found'.format(r + 2, ds_name))
                    print(dp_uid)
                    continue
                elif action.value == 'DELETE':
                    print('Graphpoints Row {}: ERROR: DELETE action to implement'.format(r + 2))
                    continue
                else:
                    print('Graphpoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))
                    continue
            # point_name = dp_result['data']['name'].split('.')[1]
        elif dp_name.value:
            # This should be a threshold
            th_uid = '{}/thresholds/{}'.format(template_uid, dp_name.value)
            th_result = template_router.callMethod('getInfo', uid=th_uid)['result']
            if not th_result['success']:
                if action.value == 'ADD':
                    print('Graphpoints Row {}: ERROR: Threshold not found'.format(r + 2, ds_name))
                    continue
                elif action.value == 'DELETE':
                    print('Graphpoints Row {}: ERROR: DELETE action to implement'.format(r + 2))
                    continue
                else:
                    print('Graphpoints Row {}: ERROR: action unknown: {}'.format(r + 2, action.value))
                    continue
            # point_name = th_result['data']['name'].split('.')[1]
        else:
            # Just about the Graphpoint
            if action.value == 'ADD':
                print('Graphpoints Row {}/{}: Datapoint or threshold is missing'.format(r+2, rows))
                continue

        # print('gp_name: {}'.format(gp_name.value))
        gp_uid = '{}/graphPoints/{}'.format(gr_uid, gp_name.value)

        gp_result = template_router.callMethod('getInfo', uid=gp_uid)['result']
        gp_present = gp_result['success'] and gp_result['data']['uid'] == gp_uid

        # print(gp_uid)
        # print(gp_result)

        if action.value == 'ADD':
            if gp_present:
                # Datapoint is present
                print('Graphpoints Row {}/{}: Graphpoint already present'.format(r + 2, rows))
            else:
                # Graphpoint doesn't exist yet
                if dp_uid:
                    gp_result = template_router.callMethod('addDataPointToGraph', dataPointUid=dp_uid,
                                                           graphUid=gr_uid)['result']
                    gp_name_default = '_'.join('{}_{}'.format(ds_name.value, dp_name.value).split('_')[1:])
                    gp_uid_default = '{}/graphPoints/{}'.format(gr_uid, gp_name_default)
                    data = {u'uid': gp_uid_default, u'newId': gp_name.value}
                    info_result = template_router.callMethod('setInfo', **data)
                    # print(info_result)

                elif th_uid:
                    gp_result = template_router.callMethod('addThresholdToGraph', thresholdUid=th_uid,
                                                           graphUid=gr_uid)['result']
                    gp_uid_default = '{}/graphPoints/{}'.format(gr_uid, dp_name.value)
                    data = {u'uid': gp_uid_default, u'newId': gp_name.value}
                    info_result = template_router.callMethod('setInfo', **data)
                    # print(info_result)

                else:
                    print('Graphpoints Row {}/{}: Datapoint or threshold not found: {}'.format(r + 2, rows, gp_name.value))
                if gp_result['success']:
                    print('Graphpoints Row {}/{}: Graphpoint/Threshold added: {}'.format(r + 2, rows, gp_name.value))
                else:
                    print('Graphpoints Row {}/{}: ERROR: Graphpoint failed: {}'.format(r + 2, rows, gp_name.value))
                    print(result)
        elif action.value == 'DELETE':
            if gp_present and not key.value:
                result = template_router.callMethod('deleteGraphPoint', uid=gp_uid)['result']
                if result['success']:
                    print('Graphpoints Row {}/{}: Graphpoint deleted'.format(r + 2, rows))
                else:
                    print('Graphpoints Row {}: ERROR: Graphpoint delete failed'.format(r + 2))
                continue
            else:
                print('Graphpoints Row {}/{}: Graphpoint already deleted'.format(r + 2, rows))
        else:
            print('Graphpoints Row {}: action unknown: {}'.format(r + 2, action.value))
            continue


        # Edit graphpoint
        if key.value:
            if action.value == 'ADD':
                # TODO: don't setInfo if value is already correct
                # TODO: Only if gp_uid is OK
                gp_data = template_router.callMethod('getInfo', uid=gp_uid)['result']['data']
                current_value =gp_data.get(key.value, '')
                if value.value == current_value:
                    print('Graphpoints Row {}/{}: Graphpoint unchanged'.format(r + 2, rows))
                    continue
                data = {u'uid': gp_uid, key.value: value.value}
                result = template_router.callMethod('setInfo', **data)['result']
                if result['success']:
                    print('Graphpoints Row {}: Graphpoint edited'.format(r + 2))
                else:
                    print('Graphpoints Row {}: Graphpoint edit failed'.format(r + 2))
                    print(result)
            elif action.value == 'DELETE':
                print('Graphpoints Row {}: DELETE action to implement'.format(r + 2))
            else:
                print('Graphpoints Row {}: action unknown: {}'.format(r + 2, action.value))

    return

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Manage local templates')
    parser.add_argument('-s', dest='environ', action='store', default='z6_prod')
    parser.add_argument('-f', dest='filename', action='store', default='local_templates.xlsx')
    options = parser.parse_args()
    environ = options.environ
    filename = options.filename

    # Routers
    dr = zenAPI.zenApiLib.zenConnector(section=environ, routerName='DeviceRouter')
    tr = zenAPI.zenApiLib.zenConnector(section=environ, routerName='TemplateRouter')

    # TODO: try
    lt = openpyxl.load_workbook(filename, read_only=True)

    # Templates
    templates_data = lt['Templates']
    process_templates(dr, tr, templates_data)
    ds_data = lt['Datasources']
    process_datasources(dr, tr, ds_data)
    dp_data = lt['Datapoints']
    process_datapoints(dr, tr, dp_data)
    th_data = lt['Thresholds']
    process_thresholds(dr, tr, th_data)
    gr_data = lt['Graphs']
    process_graphs(dr, tr, gr_data)
    gp_data = lt['Graphpoints']
    process_graphpoints(dr, tr, gp_data)


